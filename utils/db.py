import pandas as pd
from openpyxl import load_workbook
import logging

log = logging.getLogger()

class Database():
    """Database class
    """
    @staticmethod
    def read_data(filename: str):
        """Read data from Excel"""
        return pd.read_excel(filename, index_col=0)
    
    @staticmethod
    def write_user(filename: str, sheetname: str, original_data: pd.DataFrame, new_row: dict):
        # Start by opening the spreadsheet and selecting the main sheet
        workbook = load_workbook(filename)
        sheet = workbook.get_sheet_by_name(sheetname)
        num_rows = len(original_data) +2
        new_row_df = pd.DataFrame([new_row])
        for i, name in enumerate(new_row_df.columns):
            sheet.cell(row=num_rows, column=i+1).value = new_row_df.loc[0,name]
            
        # Save the spreadsheet
        workbook.save(filename = filename)
    
    @staticmethod
    def write_recommendations(filename: str, sheetname: str, data_raw: pd.DataFrame):
        data = data_raw.copy().reset_index(drop=True)
        # Start by opening the spreadsheet and selecting the main sheet
        workbook = load_workbook(filename)
        sheet2 = workbook.create_sheet(title=sheetname)
        for i, name in enumerate(data.columns):
            sheet2.cell(row=1, column=i+1).value = name
        
        for i, row in data.iterrows():
            for j, name_col in enumerate(row.index):
                # write operation perform     
                sheet2.cell(row=i+2, column=j+1).value = row[name_col]
            
        # Save the spreadsheet
        workbook.save(filename=filename)
        workbook.close()